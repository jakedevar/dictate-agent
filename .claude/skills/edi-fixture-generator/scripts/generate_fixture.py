#!/usr/bin/env python3
"""
EDI Test Fixture Generator

Generates Excel test fixtures for ClinMetric EDI assessment data.

Usage:
    generate_fixture.py <first_name> <last_name> <sessions> [options]

Options:
    --report-type TYPE    One of: both, caregiver, self-report (default: both)
    --pattern PATTERN     One of: improving, stable, variable, worsening (default: improving)
    --output DIR          Output directory (default: tests/fixtures)
    --seed N              Random seed for reproducibility

Examples:
    generate_fixture.py Hannah Barbara 50
    generate_fixture.py John Smith 10 --report-type caregiver --pattern stable
    generate_fixture.py Jane Doe 25 --pattern variable --seed 42
"""

import sys
import random
from datetime import datetime, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# Excel column headers matching Microsoft Forms EDI export format (Example_Data.xlsx)
HEADERS = [
    "Id",                                           # A (0)
    "Start time",                                   # B (1)
    "Completion time",                              # C (2)
    "Email",                                        # D (3)
    "Name",                                         # E (4)
    "Child or teen's First\xa0Name:",               # F (5) - note: \xa0 is non-breaking space
    "Child or teen's Last Name:",                   # G (6)
    "Enter child or teen's first and last initial followed by birth year (example: EP1996)",  # H (7)
    "I am the:",                                    # I (8)
    # Caregiver EDI items (columns J-V, indices 9-21)
    "How much of a problem has this been\xa0in the last 7 days?\n\nRemember to consider how the person is with others and how the person does indifferent situations and places.\n.Has explosive outbursts",
    "How much of a problem has this been\xa0in the last 7 days?\n\nRemember to consider how the person is with others and how the person does indifferent situations and places.\n.Cries or stays angry for 5 minut",
    "How much of a problem has this been\xa0in the last 7 days?\n\nRemember to consider how the person is with others and how the person does indifferent situations and places.\n.Has extreme or intense emotional",
    "How much of a problem has this been\xa0in the last 7 days?\n\nRemember to consider how the person is with others and how the person does indifferent situations and places.\n.Hard to calm him/her down when h",
    "How much of a problem has this been\xa0in the last 7 days?\n\nRemember to consider how the person is with others and how the person does indifferent situations and places.\n.Does not seem to enjoy anything",
    "How much of a problem has this been\xa0in the last 7 days?\n\nRemember to consider how the person is with others and how the person does indifferent situations and places.\n.Emotions go from 0 to 100 instan",
    "How much of a problem has this been\xa0in the last 7 days?\n\nRemember to consider how the person is with others and how the person does indifferent situations and places.\n.Has trouble calming him/herself ",
    "How much of a problem has this been\xa0in the last 7 days?\n\nRemember to consider how the person is with others and how the person does indifferent situations and places.\n.Very little makes him/her happy",
    "How much of a problem has this been\xa0in the last 7 days?\n\nRemember to consider how the person is with others and how the person does indifferent situations and places.\n.Reactions are usually more sever",
    "How much of a problem has this been\xa0in the last 7 days?\n\nRemember to consider how the person is with others and how the person does indifferent situations and places.\n.Refuses to leave the house or go",
    "How much of a problem has this been\xa0in the last 7 days?\n\nRemember to consider how the person is with others and how the person does indifferent situations and places.\n.Not responsive to praise or good",
    "How much of a problem has this been\xa0in the last 7 days?\n\nRemember to consider how the person is with others and how the person does indifferent situations and places.\n.Seems sad or unhappy",
    "How much of a problem has this been\xa0in the last 7 days?\n\nRemember to consider how the person is with others and how the person does indifferent situations and places.\n.Appears uneasy through the day",
    # Self-report EDI items (columns W-AI, indices 22-34)
    ".I had extreme or intense emotional reactions",
    ".I did not enjoy anything",
    ".I had trouble calming myself down",
    ".I felt like I was in a rage",
    ".Very few things made me happy",
    ".My reactions were usually more severe than the situation called for",
    ".I could not push myself to do what I needed to do",
    ".I had trouble feeling excited even when something good happened",
    ".I got so upset or angry that I couldn't do what I needed to do",
    ".My emotions went from calm to out of control quickly",
    ".I felt sad or unhappy",
    ".It was hard for me to cheer myself up",
    ".I felt uneasy throughout the day",
]

CAREGIVER_RESPONSES = ["Not at all", "Mild", "Moderate", "Severe", "Very Severe"]
SELF_REPORT_RESPONSES = ["Level 0", "Level 1", "Level 2", "Level 3", "Level 4"]


def get_base_level(session_num: int, total_sessions: int, pattern: str) -> tuple[float, float]:
    """
    Calculate base severity level based on session progress and pattern.
    Returns (caregiver_base, self_report_base) as floats 0-4.
    """
    progress = session_num / max(1, total_sessions - 1) if total_sessions > 1 else 0

    if pattern == "improving":
        # Start high (~3.5), end low (~1.0)
        cg_base = 3.5 - (2.5 * progress)
        sr_base = 3.0 - (2.0 * progress)
    elif pattern == "worsening":
        # Start low (~1.0), end high (~3.5)
        cg_base = 1.0 + (2.5 * progress)
        sr_base = 1.0 + (2.0 * progress)
    elif pattern == "stable":
        # Stay around moderate (~2.0-2.5)
        cg_base = 2.5
        sr_base = 2.0
    elif pattern == "variable":
        # Oscillate with some randomness
        cg_base = 2.5 + 1.5 * ((-1) ** session_num) * (0.5 + 0.5 * random.random())
        sr_base = 2.0 + 1.0 * ((-1) ** session_num) * (0.5 + 0.5 * random.random())
    else:
        cg_base = 2.5
        sr_base = 2.0

    return cg_base, sr_base


def generate_responses(base_level: float, responses: list[str], stddev: float = 0.8) -> list[str]:
    """Generate 13 EDI item responses around a base severity level."""
    result = []
    for _ in range(13):
        level = max(0, min(4, int(base_level + random.gauss(0, stddev))))
        result.append(responses[level])
    return result


def generate_fixture(
    first_name: str,
    last_name: str,
    num_sessions: int,
    report_type: str = "both",
    pattern: str = "improving",
    output_dir: str = "tests/fixtures",
    seed: int | None = None
) -> Path:
    """
    Generate an EDI test fixture Excel file.

    Args:
        first_name: Patient's first name
        last_name: Patient's last name
        num_sessions: Number of sessions to generate
        report_type: "both", "caregiver", or "self-report"
        pattern: "improving", "stable", "variable", or "worsening"
        output_dir: Directory to write the file
        seed: Random seed for reproducibility

    Returns:
        Path to the generated file
    """
    if seed is not None:
        random.seed(seed)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Write headers
    for col, header in enumerate(HEADERS, 1):
        ws.cell(row=1, column=col, value=header)

    # Generate initials + DOB code
    initials = f"{first_name[0]}{last_name[0]}0315"
    full_name = f"{first_name} {last_name}"
    start_date = datetime(2024, 1, 8)

    row = 2
    for session_num in range(num_sessions):
        session_date = start_date + timedelta(weeks=session_num)
        cg_base, sr_base = get_base_level(session_num, num_sessions, pattern)

        # Caregiver row
        if report_type in ("both", "caregiver"):
            ws.cell(row=row, column=1, value=row - 1)  # ID
            ws.cell(row=row, column=2, value=session_date)  # Start time
            ws.cell(row=row, column=3, value=session_date)  # Completion time
            ws.cell(row=row, column=4, value="caregiver@example.com")
            ws.cell(row=row, column=5, value=full_name)
            ws.cell(row=row, column=6, value=first_name)
            ws.cell(row=row, column=7, value=last_name)
            ws.cell(row=row, column=8, value=initials)
            ws.cell(row=row, column=9, value=f"a parent or caregiver of {first_name}")

            # 13 caregiver EDI items (columns J-V, 1-indexed: 10-22)
            responses = generate_responses(cg_base, CAREGIVER_RESPONSES)
            for i, resp in enumerate(responses):
                ws.cell(row=row, column=10 + i, value=resp)
            row += 1

        # Self-report row
        if report_type in ("both", "self-report"):
            ws.cell(row=row, column=1, value=row - 1)  # ID
            ws.cell(row=row, column=2, value=session_date)  # Start time
            ws.cell(row=row, column=3, value=session_date)  # Completion time
            ws.cell(row=row, column=4, value="self@example.com")
            ws.cell(row=row, column=5, value=full_name)
            ws.cell(row=row, column=6, value=first_name)
            ws.cell(row=row, column=7, value=last_name)
            ws.cell(row=row, column=8, value=initials)
            ws.cell(row=row, column=9, value="the child or teen completing this survey")

            # 13 self-report EDI items (columns W-AI, 1-indexed: 23-35)
            responses = generate_responses(sr_base, SELF_REPORT_RESPONSES, stddev=0.7)
            for i, resp in enumerate(responses):
                ws.cell(row=row, column=23 + i, value=resp)
            row += 1

    # Create output path
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    filename = f"TF_{first_name}_{last_name}.xlsx"
    file_path = output_path / filename

    wb.save(file_path)
    return file_path


def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="Generate EDI test fixture Excel files",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    %(prog)s Hannah Barbara 50
    %(prog)s John Smith 10 --report-type caregiver --pattern stable
    %(prog)s Jane Doe 25 --pattern variable --seed 42
        """
    )
    parser.add_argument("first_name", help="Patient's first name")
    parser.add_argument("last_name", help="Patient's last name")
    parser.add_argument("sessions", type=int, help="Number of sessions")
    parser.add_argument(
        "--report-type",
        choices=["both", "caregiver", "self-report"],
        default="both",
        help="Type of reports to include (default: both)"
    )
    parser.add_argument(
        "--pattern",
        choices=["improving", "stable", "variable", "worsening"],
        default="improving",
        help="Score pattern over time (default: improving)"
    )
    parser.add_argument(
        "--output",
        default="tests/fixtures",
        help="Output directory (default: tests/fixtures)"
    )
    parser.add_argument(
        "--seed",
        type=int,
        help="Random seed for reproducibility"
    )

    args = parser.parse_args()

    file_path = generate_fixture(
        first_name=args.first_name,
        last_name=args.last_name,
        num_sessions=args.sessions,
        report_type=args.report_type,
        pattern=args.pattern,
        output_dir=args.output,
        seed=args.seed
    )

    print(f"Generated: {file_path}")
    print(f"  Patient: {args.first_name} {args.last_name}")
    print(f"  Sessions: {args.sessions}")
    print(f"  Report type: {args.report_type}")
    print(f"  Pattern: {args.pattern}")


if __name__ == "__main__":
    main()
